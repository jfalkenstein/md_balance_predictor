import calendar
import os
from collections import defaultdict
from datetime import date, timedelta
from decimal import Decimal
from json import load
from logging import basicConfig, info, exception, DEBUG
from logging.handlers import RotatingFileHandler
from pathlib import Path
from typing import List, NamedTuple, Dict

import pandas as pd
from dateutil.parser import isoparse
from openpyxl import Workbook
from openpyxl.cell import Cell
from openpyxl.formatting.rule import FormatObject, ColorScaleRule
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.cell_range import CellRange
from sklearn.cluster import KMeans

from config import (
    FILE_NAME,
    REQUIRED_TXN_FREQUENCY,
    REQUIRED_TXN_PERCENT_OF_MONTHS,
    MAX_FREQUENCY_FOR_MONTHLY_RECURRANCE
)
from receiver_resources.enums import RecurrenceType, ValueConsistency
from receiver_resources.utilities import get_week_of_month, map_value_consistency, iterate_days

current_dir = Path(__file__).parent
log_file = current_dir.joinpath('logs.txt')

file_handler = RotatingFileHandler(log_file)
basicConfig(handlers=[file_handler], level=DEBUG)





def run():
    try:
        filepath = current_dir.joinpath(FILE_NAME)
        dejsonified_data = load_json_from_file(filepath)
        accounts = dejsonified_data['accounts']
        transactions = dejsonified_data['transactions']
        current_account_balance = Decimal(dejsonified_data['current_account_balance'])
        start_date = isoparse(dejsonified_data['start_date'])
        days_to_predict = dejsonified_data['days_to_predict']

        accounts_df = pd.DataFrame(accounts)
        transactions_df = create_transactions_dataframe(transactions)

        accounts_df = evaluate_account_recurrence(accounts_df, transactions_df)
        accounts_df = evaluate_average_transaction_amounts(accounts_df, transactions_df)

        range_start = start_date
        range_end = range_start + timedelta(days=days_to_predict)

        balance_predictor_df = create_balance_predictor(
            current_account_balance,
            accounts_df,
            range_start,
            range_end
        )

        predictor_path = current_dir.joinpath('predictor.xlsx')

        create_balance_predictor_xlsx(balance_predictor_df, predictor_path, current_account_balance)
        os.startfile(predictor_path)
    except Exception:
        exception("Exception encountered.")
        raise


def load_json_from_file(filepath: str) -> dict:
    info("Attempting to load json.")
    with open(filepath) as f:
        dejsonified_data = load(f)
    info("json loaded...")
    return dejsonified_data


def create_transactions_dataframe(transactions: List[dict]) -> pd.DataFrame:
    transactions_df = pd.DataFrame(transactions)
    transactions_df['value'] = -transactions_df['value'].map(lambda x: int((Decimal(x) * 100)))
    transactions_df['date'] = transactions_df['date'].map(isoparse)
    transactions_df['month'] = transactions_df['date'].map(lambda x: x.month)
    transactions_df['day_of_month'] = transactions_df['date'].map(lambda x: x.day)
    transactions_df['week_of_month'] = transactions_df['date'].map(get_week_of_month)
    transactions_df['day_of_week'] = transactions_df['date'].map(lambda x: x.weekday())
    return transactions_df


def evaluate_account_recurrence(accounts_df: pd.DataFrame, transactions_df: pd.DataFrame):
    frequency_counts_df = get_frequency_counts(transactions_df)
    accounts_df = accounts_df.merge(frequency_counts_df)
    frequent_transactions = accounts_df['frequency'] >= REQUIRED_TXN_FREQUENCY
    transactions_in_most_months = accounts_df['percent_of_months'] > REQUIRED_TXN_PERCENT_OF_MONTHS
    accounts_df = accounts_df[
        frequent_transactions & transactions_in_most_months
        ]
    categories_df = categorize_account_usage(accounts_df, transactions_df)
    return categories_df


def get_frequency_counts(transactions_df: pd.DataFrame):
    filtered_transactions = filter_out_incomplete_months(transactions_df)
    account_id_counts: pd.DataFrame = filtered_transactions.groupby('account_id').count().iloc[:, 0]
    account_id_counts.name = 'account_id_counts'
    account_id_count_df = pd.DataFrame(account_id_counts)

    month_count = len(filtered_transactions['month'].unique())

    account_id_count_df['frequency'] = account_id_count_df['account_id_counts'] / month_count
    frequency_df = account_id_count_df.reset_index()[['account_id', 'frequency']]

    unique_month_numbers_by_account = filtered_transactions.groupby('account_id').nunique()['month']
    unique_month_numbers_by_account.name = 'unique_month_count'
    unique_month_numbers_df = pd.DataFrame(
        unique_month_numbers_by_account
    ).reset_index()[['account_id', 'unique_month_count']]

    frequency_df = frequency_df.merge(unique_month_numbers_df)
    frequency_df['percent_of_months'] = frequency_df['unique_month_count'] / month_count
    return frequency_df


def filter_out_incomplete_months(transactions_df: pd.DataFrame) -> pd.DataFrame:
    transactions_for_end_of_month_transactions = transactions_df[transactions_df['week_of_month'] >= 4]
    transaction_for_start_of_month_transactions = transactions_df[transactions_df['week_of_month'] <= 2]
    start_of_month_months = transaction_for_start_of_month_transactions['month'].unique()
    end_of_month_months = transactions_for_end_of_month_transactions['month'].unique()
    complete_months_of_transactions = set(start_of_month_months).intersection(end_of_month_months)
    return transactions_df[
        transactions_df['month'].isin(complete_months_of_transactions)
    ]


def categorize_account_usage(accounts_df: pd.DataFrame, transactions_df: pd.DataFrame) -> pd.DataFrame:
    recurrency_by_account = find_recurrency(accounts_df, transactions_df)
    accounts_df = accounts_df.merge(recurrency_by_account)
    return accounts_df


def find_recurrency(accounts_df: pd.DataFrame, transactions_df: pd.DataFrame) -> pd.DataFrame:
    merged = transactions_df.merge(accounts_df, on='account_id')
    grouped_by_account_id = merged.groupby('account_id').apply(find_recurrency_on_account_transactions)
    grouped_by_account_id.index.name == 'account_id'
    grouped_by_account_id = grouped_by_account_id.reset_index()
    return grouped_by_account_id


def find_recurrency_on_account_transactions(account_transactions_df: pd.DataFrame) -> pd.Series:
    frequency: float = account_transactions_df['frequency'].values[0]
    std_on_day_of_month_by_month_is_0 = (
            account_transactions_df.groupby(['month']).std()['day_of_month'] == 0
    ).all()
    last_date_of_transaction = account_transactions_df['date'].max()
    recurring_days = None
    recurring_day_of_week = None

    if std_on_day_of_month_by_month_is_0 or frequency == 1:
        recurrance_type = RecurrenceType.ONCE_PER_MONTH
    else:
        day_of_month_recurrance_factor = account_transactions_df['day_of_month'].nunique() / max((frequency, 1))
        probably_recurrent_by_days_of_month = (
                frequency <= MAX_FREQUENCY_FOR_MONTHLY_RECURRANCE and
                day_of_month_recurrance_factor < 4
        )
        day_of_week_std = account_transactions_df.std()['day_of_week']
        recurring_day_of_week = day_of_week_std < 1

        recurrance_type = (
            RecurrenceType.BY_DAY_OF_MONTH if probably_recurrent_by_days_of_month else
            RecurrenceType.BY_DAY_OF_WEEK if recurring_day_of_week else
            RecurrenceType.RANDOM_INFREQUENT if frequency < 1 else
            RecurrenceType.RANDOM
        )

    if recurrance_type in (RecurrenceType.BY_DAY_OF_MONTH, RecurrenceType.ONCE_PER_MONTH):
        clusters = 1 if recurrance_type == RecurrenceType.ONCE_PER_MONTH else max(int(frequency), 1)
        model = KMeans(clusters)
        data = account_transactions_df[['day_of_month']]
        data['_'] = 1
        model.fit(data)
        recurring_days = pd.DataFrame(model.cluster_centers_)[0].map(round).tolist()
        recurring_day_of_week = None
    elif recurrance_type == RecurrenceType.BY_DAY_OF_WEEK:
        recurring_days = None
        recurring_day_of_week = int(account_transactions_df['day_of_week'].mode()[0])
    elif recurrance_type == RecurrenceType.RANDOM_INFREQUENT:
        recurring_days = [account_transactions_df['day_of_month'].median()]
        recurring_day_of_week = None

    return pd.Series(
        {
            'recurrance_type': recurrance_type,
            'frequency': frequency,
            'recurring_days': recurring_days,
            'recurring_day_of_week': recurring_day_of_week,
            'last_date_of_transaction': last_date_of_transaction
        },
    )


def evaluate_average_transaction_amounts(accounts_df: pd.DataFrame, transactions_df: pd.DataFrame):
    merged = transactions_df.merge(accounts_df, on='account_id')
    grouped_by_account_id = merged.groupby('account_id')
    average_value_by_account_id = grouped_by_account_id.mean()['value'].map(round)
    average_value_by_account_id.name = 'average_transaction_amount'

    earliest_date = transactions_df['date'].min()
    latest_date = transactions_df['date'].max()
    weeks = (latest_date - earliest_date).days / 7

    weekly_average = (grouped_by_account_id.sum()['value'] / weeks).map(round)
    weekly_average.name = 'weekly_average_transaction_amount'
    cv_on_values = transactions_df.groupby('account_id').apply(calculate_coefficient_of_variation)
    cv_on_values.name = 'cv_on_values'

    values_df = pd.DataFrame({
        average_value_by_account_id.name: average_value_by_account_id,
        weekly_average.name: weekly_average,
        cv_on_values.name: cv_on_values
    })
    values_df.index.name = 'account_id'
    values_df = values_df.reset_index()

    accounts_df = accounts_df.merge(values_df)
    accounts_df['value_variance'] = accounts_df['cv_on_values'].map(map_value_consistency)
    return accounts_df


def calculate_coefficient_of_variation(transactions_df: pd.DataFrame):
    if len(transactions_df) == 1 or transactions_df['value'].sum() == 0:
        return 1

    cv = transactions_df['value'].std() / abs(transactions_df['value'].mean())
    return cv


class Row:
    def __init__(self, description: str, frequency: float, amount: int, value_variance: str, recurrance_type: str, needs_attention=False):
        self.description = description
        self.frequency = frequency
        self.amount = amount
        self.value_variance = value_variance
        self.recurrance_type = recurrance_type
        self.needs_attention = needs_attention


def create_balance_predictor(current_account_balance: Decimal, accounts_df: pd.DataFrame, range_start: date,
                             range_end: date) -> pd.DataFrame:
    row_accumulator: Dict[date, List[Row]] = defaultdict(list)
    balance_start_date = range_start - timedelta(days=1)
    row_accumulator[balance_start_date].append(
        Row("BALANCE START", 0, int(current_account_balance * 100), 'N/A', 'N/A')
    )
    for _, account_row in accounts_df.iterrows():
        value_variance = account_row['value_variance']
        recurrance_type = account_row['recurrance_type']
        last_known_date = account_row['last_date_of_transaction']
        recurring_days_of_month = account_row['recurring_days']
        frequency: float = account_row['frequency']

        row = Row(
            f"{account_row['account_name']}",
            round(frequency, 3),
            account_row['average_transaction_amount'],
            value_variance.name,
            recurrance_type.name,
            needs_attention=(
                    value_variance in (ValueConsistency.VERY_VARIANT, ValueConsistency.SOME_VARIANCE) and
                    recurrance_type != RecurrenceType.RANDOM
            )
        )
        row_written = False
        if recurrance_type in (RecurrenceType.ONCE_PER_MONTH, RecurrenceType.BY_DAY_OF_MONTH):
            if not frequency.is_integer():
                row.needs_attention = True

            recurring_dates = iterate_by_recurring_days_of_month(
                last_known_date,
                frequency,
                range_start,
                range_end,
                recurring_days_of_month
            )
            for day in recurring_dates:
                row_accumulator[day].append(row)
                row_written = True
        elif recurrance_type == RecurrenceType.BY_DAY_OF_WEEK:
            day_of_week = account_row['recurring_day_of_week']
            date_iterator = iterate_weekly_recurrance_dates(
                last_known_date,
                frequency,
                range_start,
                range_end,
                int(day_of_week)
            )
            for row_date in date_iterator:
                row_accumulator[row_date].append(row)
                row_written = True
        elif recurrance_type == RecurrenceType.RANDOM_INFREQUENT:
            row.needs_attention = True
            date_iterator = iterate_days_by_frequency(
                last_known_date,
                range_start,
                range_end,
                frequency
            )
            for row_date in date_iterator:
                row_accumulator[row_date].append(row)
                row_written = True
        elif recurrance_type == RecurrenceType.RANDOM:
            row.amount = account_row['weekly_average_transaction_amount']
            for row_date in iterate_days(range_start, range_end):
                if row_date.weekday() == 0:
                    row_accumulator[row_date].append(row)
                    row_written = True

        if not row_written:
            day_after_range = range_end + timedelta(days=1)
            row.description = (
                f"{account_row['account_name']} ("
                f"NO RECORDS ADDED - "
                f"AVG. TXN: {Decimal(account_row['average_transaction_amount']) / 100} "
                f")"
            )
            row.amount = 0
            row_accumulator[day_after_range].append(row)

    balance_predictor = pd.DataFrame()
    for transaction_date, row_tuples in sorted(row_accumulator.items(), key=lambda x: x[0]):
        for row in row_tuples:
            decimal_value = Decimal(row.amount) / 100
            balance_predictor = balance_predictor.append(
                {
                    'Date': transaction_date.date(),
                    'Description': row.description,
                    'Frequency': row.frequency,
                    'Amount': decimal_value,
                    'Value Variance': row.value_variance,
                    'Recurrence Type': row.recurrance_type,
                    'Needs Attention': str(row.needs_attention)
                },
                ignore_index=True
            )

    balance_predictor = balance_predictor[[
        'Date',
        'Description',
        'Frequency',
        'Amount',
        'Value Variance',
        'Recurrence Type',
        'Needs Attention'
    ]]

    return balance_predictor


def iterate_weekly_recurrance_dates(last_known_date: date, frequency: float, start_date: date, end_date: date,
                                    day_of_week: int):
    eval_func = lambda day: day.weekday() == day_of_week
    yield from iterate_days_by_frequency(last_known_date, start_date, end_date, frequency, eval_func)


def iterate_by_recurring_days_of_month(last_known_date: date, frequency: float, start_date: date, end_date: date,
                                       recurring_days):
    if frequency.is_integer():
        for day in iterate_days(start_date, end_date):
            if day.day in recurring_days:
                yield day
    else:
        eval_func = lambda day: day.day in recurring_days
        yield from iterate_days_by_frequency(last_known_date, start_date, end_date, frequency, eval_func)


def iterate_days_by_frequency(last_known_date: date, start_date: date, end_date: date, frequency: float,
                              eval_func=lambda day: True):
    day_counter = 0
    remainder = 0
    for day in iterate_days(last_known_date, end_date):
        if date == last_known_date:
            continue
        days_in_month = calendar.monthrange(day.year, day.month)[1]
        days_per_frequency_increment = days_in_month / frequency
        day_counter += 1
        threshold = int(days_per_frequency_increment - 6 + remainder)
        if day_counter >= threshold and eval_func(day):
            remainder += (days_per_frequency_increment - day_counter) % 7
            day_counter = 0
            if start_date <= day <= end_date:
                yield day


def create_balance_predictor_xlsx(balance_predictor_df: pd.DataFrame, predictor_path: Path, current_balance):
    wb = Workbook()
    ws = wb.active
    amount_column_number = balance_predictor_df.columns.get_loc('Amount') + 1
    amount_column_letter = get_column_letter(amount_column_number)
    needs_attention_column_number = balance_predictor_df.columns.get_loc('Needs Attention') + 1
    needs_attention_column_letter = get_column_letter(needs_attention_column_number)
    balance_column_number = len(balance_predictor_df.columns) + 1
    balance_column_letter = get_column_letter(balance_column_number)
    bottom_row_number = len(balance_predictor_df) + 1
    for index, row in enumerate(dataframe_to_rows(balance_predictor_df, False)):
        row_number = index + 1
        if index == 0:  # header
            row.append('Running Balance')
        elif index == 1:  # Starting balance
            row.append(current_balance)
        else:
            formula = f'={balance_column_letter}{row_number - 1} + {amount_column_letter}{row_number}'
            row.append(formula)
        ws.append(row)

    ws.auto_filter.ref = f"A1:{get_column_letter(balance_column_number - 1)}{bottom_row_number}"
    rule = ColorScaleRule(
        start_type='percentile',
        start_color='ff0000',
        mid_type='percentile',
        mid_value=50,
        mid_color='ffff00',
        end_type='percentile',
        end_color='009900'
    )
    ws.conditional_formatting.add(f'{balance_column_letter}1:{balance_column_letter}{bottom_row_number}', rule)
    for cell in ws[f'A1:A{bottom_row_number}']:
        cell[0].number_format = 'mm/dd/yy'
    for cell in ws[f'{amount_column_letter}1:{amount_column_letter}{bottom_row_number}']:
        cell[0].number_format = '[$$-409]#,##0.00;[RED]-[$$-409]#,##0.00'
    for cell in ws[f'{balance_column_letter}1:{balance_column_letter}{bottom_row_number}']:
        cell[0].number_format = '[$$-409]#,##0.00;[RED]-[$$-409]#,##0.00'
    for cell in ws[f'{needs_attention_column_letter}1:{needs_attention_column_letter}{bottom_row_number}']:
        if cell[0].value == 'True':
            cell[0].font = Font(bold=True)
    wb.save(predictor_path)


if __name__ == '__main__':
    run()
